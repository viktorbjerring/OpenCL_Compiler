from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib.ticker

if __name__ == "__main__":
    workbook = load_workbook(filename="WorkGroups.xlsx", data_only=True)
    print(f"Generating Graphs")
    for sheet in workbook.worksheets:
        # print(sheet.title)
        name = sheet.cell(1, 1).value
        print(f'Generating graph for "{name}"')
        x_title = sheet.cell(2, 1).value
        y_title = sheet.cell(2, 3).value
        k = 3
        x_data = []
        y_data = []
        while True:
            if sheet.cell(k, 1).value:
                x_data.append(sheet.cell(k, 1).value)
                y_data.append(sheet.cell(k, 3).value)
                k += 1
                continue
            break
        fig = plt.figure()
        ax = fig.add_subplot(111)
        fig.get_axes()[0].set_xlabel(x_title)
        fig.get_axes()[0].set_ylabel(y_title)
        ax.set_xscale("log", base=2)
        ax.get_xaxis()
        ax.plot(x_data, y_data, "o")
        ax.get_xaxis().set_major_formatter(matplotlib.ticker.ScalarFormatter())
        ax.get_xaxis().set_minor_formatter(matplotlib.ticker.NullFormatter())
        # plt.xlabel(x_title)
        # plt.ylabel(y_title)
        plt.title(name)
        # plt.xscale("log", base=2)
        plt.savefig(f"{name.replace(' ', '_')}.pdf")
        plt.close()
    # Under 8192
    for sheet in workbook.worksheets:
        # print(sheet.title)
        name = f"{sheet.cell(1, 1).value} (Values of 8192 and under)"
        print(f'Generating graph for "{name}"')
        x_title = sheet.cell(2, 1).value
        y_title = sheet.cell(2, 3).value
        k = 3
        x_data = []
        y_data = []
        while True:
            if sheet.cell(k, 1).value:
                if sheet.cell(k, 1).value > 9000:
                    k += 1
                    continue
                x_data.append(sheet.cell(k, 1).value)
                y_data.append(sheet.cell(k, 3).value)
                k += 1
                continue
            break
        fig = plt.figure()
        ax = fig.add_subplot(111)
        fig.get_axes()[0].set_xlabel(x_title)
        fig.get_axes()[0].set_ylabel(y_title)
        ax.set_xscale("log", base=2)
        ax.get_xaxis()
        ax.plot(x_data, y_data, "o")
        ax.get_xaxis().set_major_formatter(matplotlib.ticker.ScalarFormatter())
        ax.get_xaxis().set_minor_formatter(matplotlib.ticker.NullFormatter())
        # plt.xlabel(x_title)
        # plt.ylabel(y_title)
        plt.title(name)
        # plt.xscale("log", base=2)
        plt.savefig(f"{name.replace(' ', '_')}.pdf")
        plt.close()

    workbook = load_workbook(filename="LexerTimings.xlsx", data_only=True)
    for sheet in workbook.worksheets:
        unitConverter = {"ns": "us", "us": "ms"}
        name = sheet.title
        unit = sheet.cell(3, 1).value
        print(f'Generating graph for "{name}"')
        x_val = [sheet.cell(2, 1).value, sheet.cell(2, 2).value]
        y_val = [sheet.cell(14, 1).value / 1000, sheet.cell(14, 2).value / 1000]
        x_val_map = [sheet.cell(2, 3).value]  # , sheet.cell(2, 4).value]
        y_val_map = [
            sheet.cell(14, 3).value / 1000
        ]  # , sheet.cell(14, 4).value / 1000]
        s1 = plt.subplot(1, 2, 1)
        s1.stem(x_val, y_val)
        # s1.tick_params("x",rotation="vertical")
        s1.set_title(name)
        plt.xticks(rotation="vertical")
        s1.set_ylabel(f"Time [{unitConverter[unit]}]")
        s2 = plt.subplot(1, 2, 2)
        s2.stem(x_val_map, y_val_map)
        new_name = name.replace("Read from", "Mapping of")
        if new_name == name:
            new_name = f"{name}\n(Memory mapping)"
        plt.xticks(rotation="vertical")
        s2.set_title(new_name)
        # s2.set_ylabel(f"Time [{unitConverter[unit]}]")
        min1, max1 = s1.get_ybound()
        min2, max2 = s2.get_ybound()
        minFinal = min(min1, min2)
        maxFinal = max(max1, max2)
        s1.set_ybound(minFinal, maxFinal)
        s2.set_ybound(minFinal, maxFinal)
        plt.savefig(f"{name.replace(' ', '_')}.pdf", bbox_inches="tight")
        plt.close()
